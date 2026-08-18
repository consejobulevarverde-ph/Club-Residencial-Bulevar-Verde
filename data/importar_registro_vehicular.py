#!/usr/bin/env python3
"""
Importa vehículos desde "BASE DE DATOS DE LOS VEHICULOS .xlsx" (registro vehicular por apartamento).

Reglas:
- Omite registros con formato de placa inválido.
- Asigna tipo (CARRO vs MOTO) según el formato de placa.
- No actualiza registros cuyas placas ya existen.
- Omite registros que no logre identificar plenamente de qué apartamento es.
- El número de torre del archivo no es confiable para matcheo — se usa solo apartamento.
- Sin --apply solamente valida y genera reportes.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import uuid
from collections import defaultdict
from dataclasses import asdict, dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable, Sequence


API_ROOT = "https://firebasedataconnect.googleapis.com/v1"
DEFAULT_PROJECT = "project-7dd6d100-d8c2-427a-a80"
DEFAULT_LOCATION = "us-east4"
DEFAULT_SERVICE = "portal-bulevar-verde"
DEFAULT_CONNECTOR = "admin"

BOGOTA = timezone(timedelta(hours=-5))
VEHICLE_NAMESPACE = uuid.UUID("25eb008f-886b-42ee-901d-b756c7b434ef")
LINK_NAMESPACE = uuid.UUID("e5b0f6c5-2e0e-44ea-8520-9185b8ef9a53")

# Regex para validar formatos de placa
REGEX_CARRO = re.compile(r"^[A-Z]{3}\d{3}$")
REGEX_MOTO = re.compile(r"^[A-Z]{3}\d{2}[A-Z]$")

# Origen del registro
ORIGEN_REGISTRO = "IMPORTACION_VIGILANCIA_20260818"


def sesion_autorizada():
    try:
        import google.auth
        from google.auth.transport.requests import AuthorizedSession
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "Falta google-auth. Ejecuta: "
            "pip install -r requirements.txt"
        ) from exc

    credenciales, _ = google.auth.default(
        scopes=[
            "https://www.googleapis.com/auth/cloud-platform"
        ]
    )
    return AuthorizedSession(credenciales)


def endpoint(
    proyecto: str,
    ubicacion: str,
    servicio: str,
    connector: str,
    metodo: str,
) -> str:
    nombre = (
        f"projects/{proyecto}/locations/{ubicacion}"
        f"/services/{servicio}/connectors/{connector}"
    )
    return f"{API_ROOT}/{nombre}:{metodo}"


def ejecutar_operacion(
    sesion,
    url: str,
    operacion: str,
    variables: dict[str, Any],
) -> dict[str, Any]:
    respuesta = sesion.post(
        url,
        json={
            "operationName": operacion,
            "variables": variables,
        },
        timeout=180,
    )
    if not respuesta.ok:
        raise RuntimeError(
            f"{operacion}: HTTP {respuesta.status_code}: "
            f"{respuesta.text[:5000]}"
        )

    resultado = respuesta.json()
    if resultado.get("errors"):
        raise RuntimeError(
            f"{operacion}: "
            + json.dumps(
                resultado["errors"],
                ensure_ascii=False,
            )
        )
    return resultado


def lotes(
    elementos: Sequence[Any],
    tamano: int,
) -> Iterable[Sequence[Any]]:
    for inicio in range(0, len(elementos), tamano):
        yield elementos[inicio:inicio + tamano]


def normalizar_placa(valor: str) -> str:
    """Normaliza placa: uppercase, solo alfanuméricos."""
    resultado = re.sub(r"[^A-Z0-9]", "", valor.upper())
    return resultado


def obtener_tipo_vehiculo(placa: str) -> str | None:
    """Infiere tipo (CARRO, MOTO) del formato de placa."""
    if REGEX_CARRO.match(placa):
        return "CARRO"
    if REGEX_MOTO.match(placa):
        return "MOTO"
    return None


def leer_excel(ruta: Path) -> list[dict[str, Any]]:
    """Lee la hoja REGISTRO VEHICULAR del XLSX."""
    try:
        import openpyxl
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "Falta openpyxl. Ejecuta: pip install openpyxl"
        ) from exc

    wb = openpyxl.load_workbook(ruta, data_only=True)

    # Busca la hoja con nombre "REGISTRO VEHICULAR " (con espacio final)
    nombre_hoja = "REGISTRO VEHICULAR "
    if nombre_hoja not in wb.sheetnames:
        raise ValueError(
            f"Hoja '{nombre_hoja}' no encontrada. "
            f"Hojas disponibles: {wb.sheetnames}"
        )

    ws = wb[nombre_hoja]

    # Detecta la última fila con datos
    ultima_fila = ws.max_row
    while ultima_fila > 0 and not any(
        ws.cell(ultima_fila, col).value
        for col in range(1, ws.max_column + 1)
    ):
        ultima_fila -= 1

    # Los datos comienzan en fila 3 (fila 2 es encabezado)
    # 5 bloques de columnas: TORRE, PLACA, APT en offsets 0, 4, 8, 12, 16
    # Cada bloque tiene 3 columnas + 1 en blanco

    entradas: list[dict[str, Any]] = []

    bloques = [
        (0, 1),   # Bloque 1: columnas A-D (TORRE en A, PLACA en B, APT en C)
        (4, 5),   # Bloque 2: columnas E-H
        (8, 9),   # Bloque 3: columnas I-L
        (12, 13), # Bloque 4: columnas M-P
        (16, 17), # Bloque 5: columnas Q-T
    ]

    for fila_num in range(3, ultima_fila + 1):
        for col_offset, _ in bloques:
            torre_col = col_offset + 1
            placa_col = col_offset + 2
            apto_col = col_offset + 3

            torre_valor = ws.cell(fila_num, torre_col).value
            placa_valor = ws.cell(fila_num, placa_col).value
            apto_valor = ws.cell(fila_num, apto_col).value

            # Si alguno está vacío, saltamos esta entrada del bloque
            if not torre_valor or not placa_valor or not apto_valor:
                continue

            entradas.append({
                "fila": fila_num,
                "bloque": (col_offset // 4) + 1,
                "torre": torre_valor,
                "placa": str(placa_valor).strip(),
                "apartamento": apto_valor,
            })

    wb.close()
    return entradas


def preparar_candidatos(
    entradas: Sequence[dict[str, Any]],
) -> tuple[
    dict[str, tuple[int, str, str]],  # placa -> (torre, apartamento, tipo)
    list[dict[str, Any]],  # omitidos_formato
    list[dict[str, Any]],  # omitidos_conflicto
]:
    """
    Procesa entradas del Excel.

    Retorna:
    - dict placa -> (torre, apartamento, tipo_vehiculo)
    - lista de omitidos por formato
    - lista de omitidos por conflicto
    """
    candidatos: dict[str, tuple[int, str, str]] = {}
    omitidos_formato: list[dict[str, Any]] = []
    omitidos_conflicto: list[dict[str, Any]] = []

    # Primera pasada: validar formatos
    registros_por_placa: dict[
        str, list[dict[str, Any]]
    ] = defaultdict(list)

    for entrada in entradas:
        placa_original = entrada["placa"]
        placa = normalizar_placa(placa_original)

        # Validar formato de placa
        tipo = obtener_tipo_vehiculo(placa)
        if not placa or not tipo:
            omitidos_formato.append({
                "torre": None,
                "apartamento": None,
                "placa": placa_original,
                "razon": f"Formato inválido: {placa_original!r}",
            })
            continue

        # Validar apartamento (debe ser parseable como int)
        apto_valor = entrada["apartamento"]
        try:
            if isinstance(apto_valor, float):
                apto_int = int(apto_valor)
            else:
                apto_int = int(str(apto_valor).strip())
            apto_normalizado = str(apto_int)
        except (ValueError, TypeError):
            omitidos_formato.append({
                "torre": None,
                "apartamento": str(apto_valor),
                "placa": placa_original,
                "razon": f"Apartamento no numérico: {apto_valor!r}",
            })
            continue

        # Parsear torre
        try:
            if isinstance(entrada["torre"], float):
                torre = int(entrada["torre"])
            else:
                torre = int(str(entrada["torre"]).strip())
        except (ValueError, TypeError):
            omitidos_formato.append({
                "torre": None,
                "apartamento": apto_normalizado,
                "placa": placa_original,
                "razon": f"Torre no numérica: {entrada['torre']!r}",
            })
            continue

        # Registrar para deduplicación por apartamento
        registros_por_placa[placa].append({
            "torre": torre,
            "apartamento": apto_normalizado,
        })

    # Segunda pasada: detectar conflictos por placa
    for placa, registros in registros_por_placa.items():
        # Agrupar por apartamento
        apartamentos_unicos = {
            r["apartamento"] for r in registros
        }

        if len(apartamentos_unicos) > 1:
            # Conflicto: misma placa en diferentes apartamentos
            omitidos_conflicto.append({
                "placa": placa,
                "torres": sorted({r["torre"] for r in registros}),
                "apartamentos": sorted(apartamentos_unicos),
                "razon": (
                    f"Placa aparece en múltiples apartamentos: "
                    f"{', '.join(apartamentos_unicos)}"
                ),
            })
            continue

        # OK: placa válida, formato OK, apartamento consistente
        r = registros[0]
        tipo = obtener_tipo_vehiculo(placa)
        candidatos[placa] = (r["torre"], r["apartamento"], tipo)

    return candidatos, omitidos_formato, omitidos_conflicto


def obtener_mapa_vehiculos(
    sesion,
    argumentos: argparse.Namespace,
) -> set[str]:
    """Obtiene el set de placas ya existentes en la base."""
    url = endpoint(
        argumentos.project,
        argumentos.location,
        argumentos.service,
        argumentos.connector,
        "impersonateQuery",
    )

    placas: set[str] = set()
    offset = 0
    limite = 500

    while True:
        resultado = ejecutar_operacion(
            sesion,
            url,
            "ObtenerMapaVehiculosImportacion",
            {
                "limit": limite,
                "offset": offset,
            },
        )
        registros = (
            resultado.get("data", {})
            .get("vehiculos", [])
        )
        for registro in registros:
            placa = registro.get("placa")
            if placa:
                placas.add(placa)

        if len(registros) < limite:
            break
        offset += limite

    return placas


def obtener_mapa_unidades(
    sesion,
    argumentos: argparse.Namespace,
) -> dict[str, list[str]]:
    """Obtiene dict: apartamento -> list[unidadId]."""
    url = endpoint(
        argumentos.project,
        argumentos.location,
        argumentos.service,
        argumentos.connector,
        "impersonateQuery",
    )

    mapa: dict[str, list[str]] = defaultdict(list)
    offset = 0
    limite = 500

    while True:
        resultado = ejecutar_operacion(
            sesion,
            url,
            "ObtenerMapaUnidadesImportacion",
            {
                "limit": limite,
                "offset": offset,
            },
        )
        registros = (
            resultado.get("data", {})
            .get("unidades", [])
        )
        for registro in registros:
            apto = registro.get("apartamento")
            id_unidad = registro.get("id")
            if apto and id_unidad:
                mapa[apto].append(id_unidad)

        if len(registros) < limite:
            break
        offset += limite

    return mapa


def importar_vehiculos(
    sesion,
    argumentos: argparse.Namespace,
    vehiculos_a_crear: Sequence[dict[str, Any]],
) -> int:
    """Importa vehículos. Retorna count."""
    url = endpoint(
        argumentos.project,
        argumentos.location,
        argumentos.service,
        argumentos.connector,
        "impersonateMutation",
    )
    importados = 0

    for numero, lote in enumerate(
        lotes(vehiculos_a_crear, argumentos.batch_size),
        start=1,
    ):
        # Filtrar campos temporales (_placa, etc.)
        datos = [
            {k: v for k, v in item.items() if not k.startswith("_")}
            for item in lote
        ]
        ejecutar_operacion(
            sesion,
            url,
            "ImportarVehiculos",
            {"datos": datos},
        )
        importados += len(lote)
        print(
            f"Vehículos lote {numero}: {len(lote)} "
            f"({importados}/{len(vehiculos_a_crear)})"
        )

    return importados


def importar_vinculos(
    sesion,
    argumentos: argparse.Namespace,
    vinculos_a_crear: Sequence[dict[str, Any]],
) -> int:
    """Importa vínculos. Retorna count."""
    url = endpoint(
        argumentos.project,
        argumentos.location,
        argumentos.service,
        argumentos.connector,
        "impersonateMutation",
    )
    importados = 0

    for numero, lote in enumerate(
        lotes(vinculos_a_crear, argumentos.batch_size),
        start=1,
    ):
        # Limpiar campos temporales (_placa, _apartamento)
        datos = [
            {k: v for k, v in item.items() if not k.startswith("_")}
            for item in lote
        ]
        ejecutar_operacion(
            sesion,
            url,
            "ImportarVinculosVehiculoUnidad",
            {"datos": datos},
        )
        importados += len(lote)
        print(
            f"Vínculos lote {numero}: {len(lote)} "
            f"({importados}/{len(vinculos_a_crear)})"
        )

    return importados


def guardar_reportes(
    carpeta: Path,
    omitidos_formato: Sequence[dict[str, Any]],
    omitidos_conflicto: Sequence[dict[str, Any]],
    omitidos_placa_existente: Sequence[dict[str, Any]],
    omitidos_unidad: Sequence[dict[str, Any]],
    vehiculos_a_crear: Sequence[dict[str, Any]],
    vinculos_a_crear: Sequence[dict[str, Any]],
    resumen: dict[str, Any],
) -> None:
    """Genera reportes en CSV y JSON."""
    carpeta.mkdir(parents=True, exist_ok=True)
    marca = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Resumen JSON
    (carpeta / f"resumen_{marca}.json").write_text(
        json.dumps(
            resumen,
            indent=2,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    # CSV: omitidos por formato inválido
    with (carpeta / f"omitidos_formato_invalido_{marca}.csv").open(
        "w",
        encoding="utf-8-sig",
        newline="",
    ) as archivo:
        escritor = csv.DictWriter(
            archivo,
            fieldnames=["apartamento", "placa", "razon"],
        )
        escritor.writeheader()
        for item in omitidos_formato:
            escritor.writerow({
                "apartamento": item.get("apartamento") or "",
                "placa": item["placa"],
                "razon": item["razon"],
            })

    # CSV: omitidos por conflicto de apartamento
    with (carpeta / f"omitidos_conflicto_apartamento_{marca}.csv").open(
        "w",
        encoding="utf-8-sig",
        newline="",
    ) as archivo:
        escritor = csv.DictWriter(
            archivo,
            fieldnames=["placa", "torres", "apartamentos", "razon"],
        )
        escritor.writeheader()
        for item in omitidos_conflicto:
            escritor.writerow({
                "placa": item["placa"],
                "torres": ";".join(str(t) for t in item["torres"]),
                "apartamentos": ";".join(item["apartamentos"]),
                "razon": item["razon"],
            })

    # CSV: omitidos por placa existente
    with (carpeta / f"omitidos_placa_existente_{marca}.csv").open(
        "w",
        encoding="utf-8-sig",
        newline="",
    ) as archivo:
        escritor = csv.DictWriter(
            archivo,
            fieldnames=["apartamento", "placa"],
        )
        escritor.writeheader()
        for item in omitidos_placa_existente:
            escritor.writerow({
                "apartamento": item["apartamento"],
                "placa": item["placa"],
            })

    # CSV: omitidos por unidad no identificada
    with (carpeta / f"omitidos_unidad_no_identificada_{marca}.csv").open(
        "w",
        encoding="utf-8-sig",
        newline="",
    ) as archivo:
        escritor = csv.DictWriter(
            archivo,
            fieldnames=["apartamento", "placa", "razon"],
        )
        escritor.writeheader()
        for item in omitidos_unidad:
            escritor.writerow({
                "apartamento": item["apartamento"],
                "placa": item["placa"],
                "razon": item["razon"],
            })

    # CSV: vehículos a crear
    modo = resumen.get("modo", "VALIDACION")
    csv_name = (
        "vehiculos_creados" if modo == "IMPORTACION"
        else "vehiculos_a_crear"
    )
    with (carpeta / f"{csv_name}_{marca}.csv").open(
        "w",
        encoding="utf-8-sig",
        newline="",
    ) as archivo:
        escritor = csv.DictWriter(
            archivo,
            fieldnames=["placa", "tipo_vehiculo", "id_destino"],
        )
        escritor.writeheader()
        for datos in vehiculos_a_crear:
            escritor.writerow({
                "placa": datos["_placa"],
                "tipo_vehiculo": datos["tipoVehiculo"],
                "id_destino": datos["id"],
            })

    # CSV: vínculos a crear
    csv_name = (
        "vinculos_creados" if modo == "IMPORTACION"
        else "vinculos_a_crear"
    )
    with (carpeta / f"{csv_name}_{marca}.csv").open(
        "w",
        encoding="utf-8-sig",
        newline="",
    ) as archivo:
        escritor = csv.DictWriter(
            archivo,
            fieldnames=[
                "placa",
                "apartamento",
                "unidad_id",
                "vinculo_id",
            ],
        )
        escritor.writeheader()
        for datos in vinculos_a_crear:
            escritor.writerow({
                "placa": datos["_placa"],
                "apartamento": datos["_apartamento"],
                "unidad_id": datos["unidadId"],
                "vinculo_id": datos["id"],
            })


def crear_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Importa vehículos desde BASE DE DATOS DE LOS VEHICULOS.xlsx"
        )
    )
    parser.add_argument(
        "--archivo",
        type=Path,
        default=Path("BASE DE DATOS DE LOS VEHICULOS .xlsx"),
    )
    parser.add_argument(
        "--apply",
        action="store_true",
    )
    parser.add_argument(
        "--batch-size",
        type=int,
        default=50,
    )
    parser.add_argument(
        "--project",
        default=DEFAULT_PROJECT,
    )
    parser.add_argument(
        "--location",
        default=DEFAULT_LOCATION,
    )
    parser.add_argument(
        "--service",
        default=DEFAULT_SERVICE,
    )
    parser.add_argument(
        "--connector",
        default=DEFAULT_CONNECTOR,
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("reportes_importacion_registro_vehicular"),
    )
    return parser


def main() -> int:
    argumentos = crear_parser().parse_args()

    try:
        if argumentos.batch_size < 1 or argumentos.batch_size > 50:
            raise ValueError(
                "--batch-size debe estar entre 1 y 50"
            )

        if not argumentos.archivo.exists():
            raise ValueError(
                f"Archivo no encontrado: {argumentos.archivo}"
            )

        # Leer y preparar candidatos
        print(f"Leyendo {argumentos.archivo}...")
        entradas = leer_excel(argumentos.archivo)
        print(f"  {len(entradas)} entradas leídas del Excel")

        candidatos, omitidos_fmt, omitidos_conf = (
            preparar_candidatos(entradas)
        )
        print(
            f"  {len(candidatos)} candidatos válidos "
            f"({len(omitidos_fmt)} formato inválido, "
            f"{len(omitidos_conf)} conflicto)"
        )

        # Consultar base de datos
        omitidos_placa_existente: list[dict[str, Any]] = []
        omitidos_unidad: list[dict[str, Any]] = []
        vehiculos_a_crear: list[dict[str, Any]] = []
        vinculos_a_crear: list[dict[str, Any]] = []

        placas_existentes: set[str] = set()
        mapa_unidades: dict[str, list[str]] = {}

        if argumentos.apply:
            sesion = sesion_autorizada()

            print("Consultando vehículos existentes...")
            placas_existentes = obtener_mapa_vehiculos(
                sesion,
                argumentos,
            )
            print(f"  {len(placas_existentes)} placas ya existen")

            print("Consultando unidades...")
            mapa_unidades = obtener_mapa_unidades(
                sesion,
                argumentos,
            )
            print(f"  {len(mapa_unidades)} apartamentos en la base")
        else:
            print("Modo validación: no se consultará la base")

        # Filtrar candidatos por reglas #3 y #4
        for placa, (torre, apto, tipo) in sorted(candidatos.items()):
            # Regla #3: Omitir si placa ya existe
            if placa in placas_existentes:
                omitidos_placa_existente.append({
                    "apartamento": apto,
                    "placa": placa,
                })
                continue

            # Regla #4: Omitir si apto no se puede identificar
            if apto not in mapa_unidades:
                omitidos_unidad.append({
                    "apartamento": apto,
                    "placa": placa,
                    "razon": "Apartamento no encontrado en la base",
                })
                continue

            unidades = mapa_unidades[apto]
            if len(unidades) > 1:
                omitidos_unidad.append({
                    "apartamento": apto,
                    "placa": placa,
                    "razon": (
                        f"Ambiguo: apartamento mapea a {len(unidades)} "
                        f"unidades (torres distintas)"
                    ),
                })
                continue

            # OK: candidato válido
            unidad_id = unidades[0]
            vehiculo_id = str(
                uuid.uuid5(VEHICLE_NAMESPACE, placa)
            )

            datos_vehiculo: dict[str, Any] = {
                "id": vehiculo_id,
                "placa": placa,
                "tipoVehiculo": tipo,
                "estadoVehiculo": "ACTIVO",
                "activo": True,
            }
            # Guardar con placa para reportes (se filtra antes de mutar)
            datos_vehiculo_con_meta = dict(datos_vehiculo)
            datos_vehiculo_con_meta["_placa"] = placa
            vehiculos_a_crear.append(datos_vehiculo_con_meta)

            link_id = str(
                uuid.uuid5(
                    LINK_NAMESPACE,
                    f"REGISTRO_VEHICULAR_2026:{placa}",
                )
            )
            ahora = datetime.now(BOGOTA).isoformat(timespec="seconds")
            datos_vinculo: dict[str, Any] = {
                "id": link_id,
                "vehiculoId": vehiculo_id,
                "unidadId": unidad_id,
                "tipoVinculo": "RESIDENTE",
                "estadoVinculo": "ACTIVO",
                "esActual": True,
                "vigenteDesde": ahora,
                "origenRegistro": ORIGEN_REGISTRO,
                "_placa": placa,  # Para reportes
                "_apartamento": apto,  # Para reportes
            }
            vinculos_a_crear.append(datos_vinculo)

        # Resumen
        print(f"\nResumen:")
        print(f"  Candidatos iniciales: {len(candidatos)}")
        print(f"  Omitidos formato inválido: {len(omitidos_fmt)}")
        print(f"  Omitidos conflicto apartamento: {len(omitidos_conf)}")
        print(f"  Omitidos placa existente: {len(omitidos_placa_existente)}")
        print(f"  Omitidos unidad no identificada: {len(omitidos_unidad)}")
        print(f"  Vehículos a crear: {len(vehiculos_a_crear)}")
        print(f"  Vínculos a crear: {len(vinculos_a_crear)}")

        vehiculos_importados = 0
        vinculos_importados = 0

        if argumentos.apply and vehiculos_a_crear:
            print("\nImportando...")
            vehiculos_importados = importar_vehiculos(
                sesion,
                argumentos,
                vehiculos_a_crear,
            )
            vinculos_importados = importar_vinculos(
                sesion,
                argumentos,
                vinculos_a_crear,
            )
            print(
                f"Importados: {vehiculos_importados} vehículos, "
                f"{vinculos_importados} vínculos"
            )
        else:
            if not argumentos.apply:
                print("Modo validación: no se importó nada")

        resumen = {
            "modo": "IMPORTACION" if argumentos.apply else "VALIDACION",
            "entradas_excel": len(entradas),
            "candidatos_validos": len(candidatos),
            "omitidos_formato_invalido": len(omitidos_fmt),
            "omitidos_conflicto_apartamento": len(omitidos_conf),
            "omitidos_placa_existente": len(omitidos_placa_existente),
            "omitidos_unidad_no_identificada": len(omitidos_unidad),
            "vehículos_a_crear": len(vehiculos_a_crear),
            "vínculos_a_crear": len(vinculos_a_crear),
            "vehículos_importados": vehiculos_importados,
            "vínculos_importados": vinculos_importados,
        }

        guardar_reportes(
            argumentos.output_dir,
            omitidos_fmt,
            omitidos_conf,
            omitidos_placa_existente,
            omitidos_unidad,
            vehiculos_a_crear,
            vinculos_a_crear,
            resumen,
        )
        print(f"\nReportes: {argumentos.output_dir.resolve()}")
        return 0

    except KeyboardInterrupt:
        print("\nProceso cancelado.", file=sys.stderr)
        return 130
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
